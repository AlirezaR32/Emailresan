import openpyxl as xl

def user_email(excel_path: str = "data/test.xlsx") -> dict:
    """Read users and emails from the first sheet of an Excel file.

    Expects names in column A and emails in column B. Skips empty rows.
    Returns a dict mapping name -> email.
    """
    wb = xl.load_workbook(excel_path)
    ws = wb.active

    max_row = ws.max_row
    users = {}
    for i in range(2, max_row + 1):
        name = ws[f"A{i}"].value
        email = ws[f"B{i}"].value
        if not name or not email:
            continue
        users[str(name).strip()] = str(email).strip()
    return users


if __name__ == "__main__":
    for a, b in user_email().items():
        print(f"{a}:{b}")


